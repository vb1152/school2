from django.contrib.auth.mixins import UserPassesTestMixin
import xlsxwriter
from django.http import HttpResponse
from openpyxl import load_workbook
from .models import Student, MyUser, UsersData
from .forms import UploadExcelFileForm
import datetime


def read_excel_with_students(request_file):
    '''Function to read data from excel file uploaded from a user
    and save data to a database'''
    wb = load_workbook(filename=request_file, data_only=True)
    sheet_ranges = wb.active
    sheet_ranges.max_column
    sheet_ranges.max_row
    stud_data = list(sheet_ranges.values)
    students_objects_list = []

    for row in stud_data[1:]:

        if Student.objects.filter(first_name=row[1], last_name=row[3], date_of_birth=row[5]).exists():
            continue
        else:
            student = Student(
                school_id=row[0],
                first_name=row[1],
                middle_name=row[2],
                last_name=row[3],
                preferred_name=row[4],
                date_of_birth=row[5],
                birth_order_in_class=int(row[6]),
                birth_order_in_family=int(row[7]),
                # add gender to db or empty value if not Male or Femail are actual data in excel.
                gender='M' if row[8] == 'Male' else (
                    'F' if row[8] == 'Female' else 'smt'),
                cur_grade=int(row[9]),
                grad_year=row[10],
                email=row[11],
                home_lang=row[12],
                date_join=row[13],
                entrygrades=row[14],
                teacher=get_teacher(row[15])
            )

        students_objects_list.append(student)

    new_sudents = Student.objects.bulk_create(
        students_objects_list)  # , ignore_conflicts=True)

    return (len(new_sudents), 'Success!')


def read_excel_save_users(request):
    form = UploadExcelFileForm(request.POST, request.FILES)
    if form.is_valid():
        request_file = request.FILES['file']
        wb = load_workbook(filename=request_file, data_only=True)
        sheet_ranges = wb.active
        sheet_ranges.max_column
        sheet_ranges.max_row
        user_data = list(sheet_ranges.values)
        user_counter = 0
        for row in user_data[1:]:
            if MyUser.objects.filter(username=row[1]).exists():
                continue
            else:
                user = MyUser.objects.create_user(
                    username=row[1],
                    password=row[1],
                    first_name=row[3],
                    last_name=row[4],
                    email=row[6])

                if row[5] == 'teacher':
                    user.is_teacher = True
                elif row[5] == 'sst':
                    user.is_sst = True
                elif row[5] == 'counselor':
                    user.is_conselor = True
                user.save()

                user_data = UsersData.objects.create(
                    user=user,
                    person_id=row[0],
                    grades=row[2])
                user_data.save()
                user_counter += 1
        return user_counter


def get_teacher(teacher_data) -> object:
    '''Find teacher instanse in data base.
    If not - return None
    -----------
    Return MyUser object (teacher == True)
    '''
    firstname, lastname = teacher_data.split()
    try:
        teach_inst = MyUser.objects.get(
            first_name=firstname,
            last_name=lastname,
            is_teacher=True)
        return teach_inst
    except MyUser.DoesNotExist:
        return None


def calculate_age(date_of_birth):
    '''Calculate students age form date of birth'''
    from datetime import datetime
    from dateutil import relativedelta

    now = datetime.now()
    # convert string to date object
    start_date = datetime.strptime(date_of_birth, "%Y-%m-%d")
    # Get the relativedelta between two dates
    delta = relativedelta.relativedelta(now, start_date)
    return (delta.years, delta.months)


def create_sample_excel_users(request):
    '''Create sample Excel file to upload users'''
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Sample_users.xlsx'

    workbook = xlsxwriter.Workbook(response)  # , {'in_memory': True})
    worksheet = workbook.add_worksheet()
    worksheet.set_column(0, 7, 20)
    cell_format = workbook.add_format({'bold': True})
    cell_format.set_bg_color('gray')

    worksheet.write('A1', 'PersonId', cell_format)
    worksheet.write('A2', '1235')
    worksheet.write('A3', '235')

    worksheet.write('B1', 'User Name', cell_format)
    worksheet.write('B2', 'John_Doe')
    worksheet.write('B3', 'James_Bond')

    worksheet.write('C1', 'Grades', cell_format)
    worksheet.write('C2', '2,3,5')
    worksheet.write('C3', 'None')

    worksheet.write('D1', 'First Name', cell_format)
    worksheet.write('D2', 'John')
    worksheet.write('D3', 'James')

    worksheet.write('E1', 'Last Name', cell_format)
    worksheet.write('E2', 'Doe')
    worksheet.write('E3', 'Bond')

    worksheet.write('F1', 'User Role', cell_format)
    worksheet.write('F2', 'teacher')
    worksheet.write('F3', 'sst')

    worksheet.write('G1', 'Email', cell_format)
    worksheet.write('G2', 'john_doe@school.com')
    worksheet.write('G3', 'james_bond@school.com')

    workbook.close()

    return response


def create_sample_excel_students(request):
    '''Create sample Excel file to upload students'''
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Sample_students.xlsx'
    workbook = xlsxwriter.Workbook(response)  # , {'in_memory': True})
    worksheet = workbook.add_worksheet()
    worksheet.set_column(0, 15, 10)
    cell_format = workbook.add_format({'bold': True})
    cell_format.set_bg_color('gray')

    worksheet.write('A1', 'Id', cell_format)
    worksheet.write('A2', 58418)
    worksheet.write('A3', 39430)

    worksheet.write('B1', 'First Name', cell_format)
    worksheet.write('B2', 'Webb')
    worksheet.write('B3', 'Shaw')

    worksheet.write('C1', 'Middle Name', cell_format)
    worksheet.write('C2', 'None')
    worksheet.write('C3', 'Alex')

    worksheet.write('D1', 'Last Name', cell_format)
    worksheet.write('D2', 'Cindy')
    worksheet.write('D3', 'Rodolfo')

    worksheet.write('E1', 'Preferred Name', cell_format)
    worksheet.write('E2', 'None')
    worksheet.write('E3', 'Alex')

    date_format = workbook.add_format({'num_format': 'dd.mm.yyyy'})
    date_time_1 = datetime.datetime.strptime('2002.06.21', '%Y.%m.%d')
    date_time_2 = datetime.datetime.strptime('2001.05.06', '%Y.%m.%d')
    worksheet.write('F1', 'DOB', cell_format)
    worksheet.write('F2', date_time_1, date_format)
    worksheet.write('F3', date_time_2, date_format)

    worksheet.write('G1', 'Birth order in class:', cell_format)
    worksheet.write('G2', 5)
    worksheet.write('G3', 6)

    worksheet.write('H1', 'Birth order in family:', cell_format)
    worksheet.write('H2', 2)
    worksheet.write('H3', 1)

    worksheet.write('I1', 'Gender', cell_format)
    worksheet.write('I2', 'Male')
    worksheet.write('I3', 'Female')

    worksheet.write('J1', 'Current Grade', cell_format)
    worksheet.write('J2', 12)
    worksheet.write('J3', 12)

    worksheet.write('K1', 'Graduation Year', cell_format)
    worksheet.write('K2', 2020)
    worksheet.write('K3', 2020)

    worksheet.write('L1', 'Email', cell_format)
    worksheet.write('L2', 'None')
    worksheet.write('L3', 'None')

    worksheet.write('M1', 'Home Language', cell_format)
    worksheet.write('M2', 'English')
    worksheet.write('M3', 'English')

    date_time_3 = datetime.datetime.strptime('2015.09.01', '%Y.%m.%d')
    date_time_4 = datetime.datetime.strptime('2013.09.03', '%Y.%m.%d')

    worksheet.write('N1', 'Date of joining', cell_format)
    worksheet.write('N2', date_time_3, date_format)
    worksheet.write('N3', date_time_4, date_format)

    worksheet.write('O1', 'Entrygrades', cell_format)
    worksheet.write('O2', 5)
    worksheet.write('O3', 5)

    worksheet.write('P1', 'Teacher', cell_format)
    worksheet.write('P2', 'John Doe')
    worksheet.write('P3', 'Amy Doe')

    workbook.close()

    return response


def teacher_check(user):
    '''Check if user is a teacher for user_passes_test decorator'''
    return user.is_teacher


def sst_check(user):
    '''Check if user is a sst for user_passes_test decorator'''
    return user.is_sst


def staff_check(user):
    '''Check if user is a sst for user_passes_test decorator'''
    return user.is_staff


class TeacherCheckMixin(UserPassesTestMixin):
    '''Check if user us a teacher'''

    def test_func(self):
        return self.request.user.is_teacher


class SstCheckMixin(UserPassesTestMixin):
    '''Check if user is a sst for user_passes_test decorator'''

    def test_func(self):
        return self.request.user.is_sst


class StaffCheckMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff
